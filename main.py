import os
import re
import shutil
import tempfile
import time
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager


# Scrolls the div or container
def scroll_job_list_container(driver, scrollable_container, steps=10, delay=0.5):
    for _ in range(steps):
        driver.execute_script("arguments[0].scrollBy(0, 300);", scrollable_container)
        time.sleep(delay)


# Function that filter the qualification from the description
def extract_qualifications(desc_text):
    patterns = [
        r"(Qualifications|Requirements|What you[’'\']ll need|"
        r"Preferred Qualifications|Skills Required|What We’re Looking|What we're looking for|"
        r"We're looking for people who have:|What You’ll Bring|Who You Are|"
        r"What You Bring|Basic Qualifications|Must Have|You Should Have|"
        r"Desired Skills|Preferred Experience|The Ideal Candidate)"
        r"[\s\S]{0,2000}"
    ]

    for pattern in patterns:
        match = re.search(pattern, desc_text, re.IGNORECASE)
        if match:
            qualification_text = match.group()
            break
    else:
        return "N/A"

    lines = qualification_text.splitlines()
    filtered = []
    for line in lines:
        line = line.strip()
        if line:
            filtered.append(line)

    return "\n".join(filtered)


def Scraping_job_Details(driver, job_links, all_jobs_data):
    for link in job_links:
        try:
            driver.get(link)
            time.sleep(5)
            soup = BeautifulSoup(driver.page_source, 'lxml')

            # Title
            try:
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'inline')))
                job_title = soup.find('h1', class_='inline').text.strip()
            except Exception as e:
                print(f"Error fetching job title: {e}")
                job_title = "N/A"

            # Company
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CLASS_NAME, 'job-details-jobs-unified-top-card__company-name')))
                job_company = soup.find('div', class_='job-details-jobs-unified-top-card__company-name').text.strip()
            except Exception as e:
                print(f"Error fetching company: {e}")
                job_company = "N/A"

            # Location
            try:
                WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                    (By.CLASS_NAME, 'job-details-jobs-unified-top-card__tertiary-description-container')))
                job_location = soup.find('div',
                                         class_='job-details-jobs-unified-top-card__tertiary-description-container').text.strip().split(
                    '·')[0]
            except Exception as e:
                print(f"Error fetching location: {e}")
                job_location = "N/A"

            # Job Type & Salary
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CLASS_NAME, 'job-details-fit-level-preferences')))
                desc_abt_job = soup.find('div', class_='job-details-fit-level-preferences').get_text(strip=True)
                cleaned = re.sub(r'\$[\d,]+K/yr\s*-\s*\$[\d,]+K/yr', '', desc_abt_job).strip()
                job_type = ' | '.join(re.findall(r'[A-Z][a-z]+(?:-[a-z]+)?', cleaned))
                salary_match = re.search(r'\$[\d,]+K/yr\s*-\s*\$[\d,]+K/yr', desc_abt_job)
                job_salary = salary_match.group() if salary_match else 'N/A'
            except Exception as e:
                print(f"Error fetching job type/salary: {e}")
                job_type = "N/A"
                job_salary = "N/A"

            # Description & Qualifications
            try:
                scrollable_container = WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.CLASS_NAME, 'jobs-semantic-search-job-details-wrapper'))
                )
                scroll_job_list_container(driver, scrollable_container)
                job_description = soup.find('div', class_='jobs-description-content__text--stretch').get_text(
                    strip=True)
                job_qualification = extract_qualifications(job_description)
            except Exception as e:
                print(f"Error fetching qualifications: {e}")
                job_qualification = "N/A"

            # Append data
            all_jobs_data.append({
                "Job Title": job_title,
                "Company": job_company,
                "Location": job_location,
                "Job Type": job_type,
                "Salary": job_salary,
                "Link": link,
                "Raw Qualifications": job_qualification
            })

        except Exception as e:
            print(None)


def Scraping_Links():
    job_search = ['Frontend Developer', 'Backend Developer', 'QN Writer']
    # Holds all jobs details
    all_jobs_data = []

    # Function that retries chrome session setup when fails
    def setup_chrome_driver(retries=5, delay=3):
        options = Options()
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option("useAutomationExtension", False)
        options.add_argument('--guest')
        attempt = 0
        while attempt < retries:
            try:
                driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
                print("ChromeDriver setup succeeded.")
                return driver
            except OSError:
                attempt += 1
                time.sleep(delay)

        print("Failed to set up ChromeDriver after multiple attempts.")
        return None

    temp_profile = tempfile.mkdtemp()
    shutil.rmtree("C:/temp/chrome-temp", ignore_errors=True)
    os.makedirs("C:/temp/chrome-temp", exist_ok=True)

    # LinkedIn Homepage Links
    linkedIn = 'https://www.linkedin.com/login'
    driver = setup_chrome_driver()
    driver.get(linkedIn)
    email = os.environ['email_id']
    password = os.environ['password']
    print(email, password, type(password))
    # Login using email and password
    email_entry = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "username")))
    email_entry.send_keys('synthshorts06@gmail.com')
    password_entry = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "password")))
    password_entry.send_keys('kama2006')
    sign_in = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@type="submit"]')))
    sign_in.click()

    for job in job_search:

        # Search the specific job in the homepage
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="global-nav-search"]/div/button'))).click()
        search_job = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="global-nav-typeahead"]/input')))
        search_job.send_keys(f'{job}, United States, Remote', Keys.ENTER)

        # Filters only job posts
        filter_job = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="search-reusables__filters-bar"]/ul/li[1]/button')))
        filter_job.click()

        next_page = True
        page_number = 1

        # Loops till last page
        while next_page:

            # Scrolls the job list container (div)
            scrollable_container = WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="main"]/div/div[2]/div[1]/ul')))
            scroll_job_list_container(driver, scrollable_container, steps=60, delay=0.2)

            linked_page_source = driver.page_source
            soup = BeautifulSoup(linked_page_source, 'lxml')

            # Scraping the links from the soup
            job_links = []
            for a_tag in soup.find_all("a", class_="job-card-job-posting-card-wrapper__card-link"):
                href = a_tag.get("href")
                if href:
                    job_links.append(href)

            # Function that scrapes the job details from the links
            Scraping_job_Details(driver, job_links, all_jobs_data)

            # Using pandas, saving the details in the excelsheet at each page
            df = pd.DataFrame(all_jobs_data)
            excel_path = "linkedin_jobs.xlsx"
            df.to_excel(excel_path, index=False)

            wb = load_workbook(excel_path)
            ws = wb.active

            for column_cells in ws.columns:
                max_length = 0
                column = column_cells[0].column
                column_letter = get_column_letter(column)

                for cell in column_cells:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = max_length + 2

            wb.save(excel_path)
            print(f"Saved page {page_number} results to linkedin_jobs_{job.replace(' ', '_')}_page_{page_number}.xlsx")

            # Goes to the next page
            next_page = soup.find(name='button', class_='jobs-search-pagination__button--next')
            if next_page and page_number < 3:
                next_page = True
                next_page_button = driver.find_element(By.CLASS_NAME, 'jobs-search-pagination__button--next')
                next_page_button.click()
                page_number += 1
            else:
                next_page = False
                break


# Scraping the links of the jobs from LinkedIn
Scraping_Links()
